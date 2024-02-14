import os
from openpyxl import Workbook, load_workbook

class ranks:
    def __init__(self, score):
        self.score = score

    def high_score(self):
        file_path = 'ranks.xlsx'

        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.append(['High score'])
            wb.save(file_path)

        wb = load_workbook(file_path)
        ws = wb.active
        scores = [cell.value for cell in ws['A'][1:]]

        if self.score > 0 and self.score not in scores:
            ws.insert_rows(2)
            ws['A2'] = self.score

        scores.sort(reverse=True)
        scores = scores[:10]
        wb.save(file_path)

        return scores
